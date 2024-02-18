# -*- coding = utf-8 -*-
# @Time : 2022/4/18 11:04
# @Author : 从小白出发
# @File : test.py
# @Software : PyCharm
import pathlib
import re
import time
import openpyxl
import requests
from lxml import etree

url = 'https://www.2bulu.com/track/track_search_result.htm?sortType=0'
head = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) '
                  'Chrome/121.0.0.0 Safari/537.36 Edg/121.0.0.0', }
# # i 代表 pageNumber,从第1页到x页
for i in range(1,10):
    #设置定时3秒访问一次,防止被封ip，实测并不需要,该网页没有针对访问频率的反爬
    time.sleep(3)
    # 荷载 ,即搜索数据
    data = {'key': '中央大街', 'pageNumber': i, 'areaId': '4346', 'parentId': '4345'}
    cook = {'webSessionId': 'C1FB580D037ACA2BF613AB5C2998FAF5-n2',
            'login_info_code': 'BLIsafjcr1X5sCD0SwwSWA==',
            'Hm_lvt_bed4e06c9c336e6fec0b7611aeefdc73': '1708053818,1708141238',
            'JSESSIONID': 'C1FB580D037ACA2BF613AB5C2998FAF5-n2'
            }
    #发送post请求
    resp = requests.post(url=url, headers=head, cookies=cook, data=data, timeout=1)
    #服务端返回状态码200代表正常
    if resp.status_code == 200:
        h = etree.HTML(resp.text)

        # 提取数据
        # 因为每一页仅展示10条数据 因此我们从1到10,j代表每一条
        for j in range(1, 11):
            # 遍历j并放到div中
            # 提取下一级链接, a是超链接, @href可以获取到url
            semi_link = h.xpath('//div[' + str(j) + ']//a/@href')
            link = 'https://www.2bulu.com' + semi_link[0]
            # 提取出行方式, text()可以获取文本内容
            way = h.xpath('//div[' + str(j) + ']//p/span[1]/text()')
            # 去除数据头和尾部的空格(\t\n)
            way = way[0].strip()
            # 提取标题
            title = h.xpath('//div[' + str(j) + ']//p[1]/text()')
            title = title[1].strip()
            # 提取出行距离
            distance = h.xpath('//div[' + str(j) + ']//li[1]/span[1]/text()')
            distance = distance[0].strip()
            # 去除数据中间的空格(\t\n)
            distance = re.sub('\t', '', distance)
            distance = re.sub('\n', '', distance)
            # 提取作者名称
            author = h.xpath('//div[' + str(j) + ']//li[1]/span[3]/a/text()')
            author = author[0].strip()
            # 提取起点和终点
            travel = h.xpath('//div[' + str(j) + ']//li[2]/text()')
            travel = travel[0].strip()
            travel = re.sub('\t', '', travel)
            travel = re.sub('\n', '', travel)
            # 提取出行时间
            tim = h.xpath('//div[' + str(j) + ']//li[1]/span[3]/text()')
            # 该网页中有的用户名存在特殊字符,比如包含'<'符号,在xpath中被误认为html元素的一部分,因此提取会出现问题
            # 本应该提取的作者名称, 实际提取成了出行时间
            # 但是在提取出行时间的时候,因为提取不到数据,因此报越界的错误
            # 所以这里做一个try catch操作,如果越界,则把作者名称赋值给出行时间,同时将作者名称赋值为'未知字符'
            try:
                tim = tim[1].strip()
            except IndexError as e:
                tim = author
                author = '未知字符'
            # 打印结果
            print(str(i - 1) + str(j - 1), link, way, title, distance, author, tim, travel)

            # 数据存储到excel
            if not pathlib.Path('myexcel.xlsx').exists():
                # 无则创建
                book = openpyxl.Workbook()
            else:
                # 有则打开
                book = openpyxl.load_workbook('myexcel.xlsx')
            # 激活当前sheet
            sheet = book.active
            # 行号
            row = (i-1)*10+j
            # 向单元格插入数据
            sheet.cell(row, 1).value = row
            sheet.cell(row, 2).value = link
            sheet.cell(row, 3).value = way
            sheet.cell(row, 4).value = title
            sheet.cell(row, 5).value = distance
            sheet.cell(row, 6).value = author
            sheet.cell(row, 7).value = tim
            sheet.cell(row, 8).value = travel
            # 保存文件
            book.save("myexcel.xlsx")
